'''
Table_dumper module creates a group of csv files out of the value table file
with the intention to prepare them for further processing.
'''
# import unicodecsv as csv
import csv
from openpyxl import load_workbook


class TableDumper:
    def __init__(self, path):
        self.path = path
        # open docuement
        self.values_table = load_workbook(
            self.path, read_only=True, data_only=True,)
        # get all sheetnames
        self.sheets = self.values_table.sheetnames
        # discard all sheets that do not contain data
        self.documentation_sheets = ['PORTADA', 'CONTENIDO', 'FIRMAS(1)'
                                     'Definiciones', 'DEF. DE ROSS-HEIDECKE']

    def dump(self):
        # dump all valuable sheets into csv files,
        for sheet in self.sheets:
            if sheet not in self.documentation_sheets:
                ws = self.values_table[sheet]
                print("sheet: \"{0}\"\nrows: {1}\ncols: {2}".format(
                    ws.title, ws.max_row, ws.max_column))
                with open('SunFlwr/res/{}.csv'.format(sheet),
                          'w', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    for row in ws.iter_rows():
                        row_list = []
                        [row_list.append(cell.value) for cell in row]
                        writer.writerow(row_list)
                        if not all(item is None for item in row_list):
                            print(row_list)
                file.close()
                print('---------------------------------')


if __name__ == '__main__':
    td = TableDumper('SunFlwr/res/tabla_valores_2019.xlsx')
    td.dump()
