'''
CellMaker module.
This module is the low-level formatter of the excel evaluation generator.
this class is the basic contruction for  doing cells  which will include:
    - cells that only contain values and a location.
    - cells that need ot be merged with more adjecent tables.
'''
from openpyxl import Workbook
from openpyxl.drawing.image import Image


def add_cell(ws, x, y, text, dx=0, dy=0):
    '''
    create cells with predifinied coodenates(cartesan) and size(merged cells).
    ws = worksheet
    x = horizontal axis (letters)
    y = vertical axis (numbers)
    length = lenght of the cell if happens to merge.
    '''
    if dx or dy > 0:
        ws.merge_cells(merge_coord_range(x, y, dx, dy))
    ws[gen_coord(x, y)] = text


def gen_coord(x, y):
    """
    coverts number to cordenates to excel format coordinates.
    in case a cells is position after the z letter the cell position will be
    looped back to the first cell in the x axis.
    """
    if x > 26:
        print("range too big; out of bounds; cell position overflow.")
    return '{0}{1}'.format(chr((x % 26)+64), y)


def merge_coord_range(x, y, dx=0, dy=0):
    '''Genrates a range of two excel coordinate points.'''
    return '{0}:{1}'.format(gen_coord(x, y), gen_coord(x+dx, y+dy))


def add_image(ws, x, y, path, dx=0, dy=0):
    '''Add an image and anchors it to the specified columns
    # TODO: add the resizing property'''
    img = Image(path)
    ws.add_image(img, gen_coord(x, y))
    pass


if __name__ == '__main__':
    print(gen_coord(1, 2))
    print(merge_coord_range(1, 2, 3))
    print(merge_coord_range(1, 2, 3, 4))
    wb = Workbook()
    ws = wb.active
    for i in range(1, 10):
        add_cell(ws, 1, i, i, dx=i)

    add_cell(ws, 27, 15, 'something', dx=2)
    img_path = 'Sunflwr/res/64978599_397524490970857_7645817694196858880_n.jpg'
    img = Image(img_path)
    ws.add_image(img, 'F20')
    wb.save('tests/sheets/loop.xlsx')
